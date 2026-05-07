#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
未发货订单统计 - 数据处理脚本
功能：基于客审状态和发货时间类型生成正确的透视表和统计表
输出：包含"Sheet1_处理后"、"透视表"、"统计表"三个工作表的Excel文件
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys


def process_data(input_file, output_dir):
    """
    处理未发货订单数据 - 使用正确的统计逻辑
    
    Args:
        input_file: 原始Excel文件路径
        output_dir: 输出目录路径
    """
    print("=" * 60)
    print("未发货订单统计 - 数据处理（已修复版本）")
    print("=" * 60)
    
    # 获取当天日期
    today = datetime.now().strftime("%Y%m%d")
    
    # 生成输出文件路径
    OUTPUT_FILE = os.path.join(output_dir, '发货计划统计结果.xlsx')
    
    print(f"\n输入文件: {input_file}")
    print(f"输出文件: {OUTPUT_FILE}")
    
    # ========== 步骤1: 读取数据 ==========
    print("\n【步骤1】读取Excel数据...")
    try:
        df = pd.read_excel(input_file, sheet_name="Sheet1")
        print(f"  原始数据: {df.shape[0]} 行, {df.shape[1]} 列")
    except Exception as e:
        print(f"  读取失败: {e}")
        # 尝试不指定sheet_name
        df = pd.read_excel(input_file)
        print(f"  自动检测sheet，读取: {df.shape[0]} 行, {df.shape[1]} 列")
    
    # ========== 步骤2: 添加"客审状态"列 ==========
    # 规则1: "标记"列含有"客审"或"货审"则为"已客审"，否则"未客审"
    print("\n【步骤2】添加'客审状态'列...")
    def get_客审状态(mark):
        if pd.isna(mark):
            return "未客审"
        mark_str = str(mark)
        if "客审" in mark_str or "货审" in mark_str:
            return "已客审"
        return "未客审"
    
    if "标记" in df.columns:
        df["客审状态"] = df["标记"].apply(get_客审状态)
        已客审_count = (df["客审状态"] == "已客审").sum()
        未客审_count = (df["客审状态"] == "未客审").sum()
        print(f"  已客审: {已客审_count}, 未客审: {未客审_count}")
    else:
        df["客审状态"] = "未客审"
        print("  警告: 未找到'标记'列，默认全部为'未客审'")
    
    # ========== 步骤3: 添加"发货时间类型"列 ==========
    # 规则2: "订单-发货通知状态"为"加急发货"或"指定时间发货"→"指定时间"
    #         含有"有货就发"或"等通知"→"等通知"
    #         空白→空白
    print("\n【步骤3】添加'发货时间类型'列...")
    def get_发货时间类型(status):
        if pd.isna(status) or str(status).strip() == "":
            return ""
        status_str = str(status).strip()
        if status_str in ["加急发货", "指定时间发货"]:
            return "指定时间"
        if "有货就发" in status_str or "等通知" in status_str:
            return "等通知"
        return ""
    
    if "订单-发货通知状态" in df.columns:
        df["发货时间类型"] = df["订单-发货通知状态"].apply(get_发货时间类型)
        等通知_count = (df["发货时间类型"] == "等通知").sum()
        指定时间_count = (df["发货时间类型"] == "指定时间").sum()
        print(f"  等通知: {等通知_count}, 指定时间: {指定时间_count}")
    else:
        df["发货时间类型"] = ""
        print("  警告: 未找到'订单-发货通知状态'列")
    
    # ========== 步骤4: 格式化日期 ==========
    # 规则3: "订单-人工承诺发货时间"改为"YYYY年MM月"格式
    print("\n【步骤4】格式化日期为'YYYY年MM月'...")
    def format_time(t):
        if pd.isna(t) or str(t).strip() == "":
            return ""
        try:
            if isinstance(t, str):
                dt = pd.to_datetime(t)
            else:
                dt = t
            return dt.strftime("%Y年%m月")
        except:
            return ""
    
    date_col = None
    for col in ["订单-人工承诺发货时间", "承诺发货时间", "发货时间", "付款时间"]:
        if col in df.columns:
            date_col = col
            break
    
    if date_col:
        df["订单-人工承诺发货时间_格式"] = df[date_col].apply(format_time)
        print(f"  使用日期列: {date_col}")
    else:
        df["订单-人工承诺发货时间_格式"] = ""
        print("  警告: 未找到日期列")
    
    # ========== 步骤5: 添加"状态合并"列 ==========
    # 规则4: "客审状态" + "-" + "发货时间类型"
    print("\n【步骤5】添加'状态合并'列...")
    df["状态合并"] = df["客审状态"] + "-" + df["发货时间类型"]
    df["状态合并"] = df["状态合并"].str.replace(r"-$", "", regex=True)
    
    # ========== 步骤6: 添加"统计项目"列 ==========
    # 规则5: 根据"状态合并"判断统计项目
    #   - 含"未客审" → "未客审"
    #   - 含"已客审-指定时间" → "订单-人工承诺发货时间"格式日期
    #   - 含"已客审-等通知" → "等通知"
    #   - 发货时间类型为空 → "未客审"
    print("\n【步骤6】添加'统计项目'列...")
    def get_统计项目(row):
        状态合并 = str(row["状态合并"])
        发货时间类型 = str(row["发货时间类型"])
        
        if "未客审" in 状态合并:
            return "未客审"
        elif "已客审-指定时间" in 状态合并:
            return row["订单-人工承诺发货时间_格式"] or "未客审"
        elif "已客审-等通知" in 状态合并:
            return "等通知"
        else:
            return "未客审"
    
    df["统计项目"] = df.apply(get_统计项目, axis=1)
    print(f"  统计项目分布:\n{df['统计项目'].value_counts().to_string()}")
    
    # ========== 步骤7: 创建透视表（自然顺序，不排序） ==========
    # 规则6: 透视表只做数据聚合，列顺序按自然排列
    print("\n【步骤7】创建透视表（自然顺序）...")
    
    # 查找销售渠道列
    channel_col = None
    for col in ["销售渠道", "渠道", "来源", "店铺"]:
        if col in df.columns:
            channel_col = col
            break
    
    if channel_col:
        df_filtered = df[df[channel_col].notna() & (df[channel_col] != "")]
        print(f"  使用渠道列: {channel_col}, 过滤后数据量: {len(df_filtered)}")
    else:
        df["销售渠道"] = "全部渠道"
        channel_col = "销售渠道"
        df_filtered = df
        print(f"  创建虚拟渠道列")
    
    # 查找金额列
    amount_col = None
    for col in ["应收合计", "实付金额", "金额", "货款", "订单金额"]:
        if col in df.columns:
            amount_col = col
            break
    
    if not amount_col:
        amount_col = "应收合计"
        df[amount_col] = 0
        print("  警告: 未找到金额列！")
    else:
        print(f"  使用金额列: {amount_col}")
    
    pivot = pd.pivot_table(
        df_filtered,
        values=amount_col,
        index=channel_col,
        columns="统计项目",
        aggfunc="sum",
        fill_value=0,
        margins=True,
        margins_name="合计"
    )
    
    print(f"  透视表形状: {pivot.shape}")
    print(f"  透视表列顺序（自然）: {list(pivot.columns)}")
    
    # ========== 步骤8: 创建统计表（排序+格式化） ==========
    # 规则7: 统计表在透视表基础上做排序和格式化
    #   列顺序：销售渠道 | 日期从小到大 | 等通知 | 未客审 | 合计
    print("\n【步骤8】创建统计表（排序+格式化）...")
    
    def get_sort_order(x):
        """列排序函数：日期(小)→等通知→未客审→合计(大)"""
        if x == "合计":
            return (9999, "")
        if x == "未客审":
            return (2, "未客审")
        if x == "等通知":
            return (1, "等通知")
        try:
            dt = pd.to_datetime(x.replace("年", "-").replace("月", ""))
            return (0, dt)
        except:
            return (3, x)
    
    # 排序列
    all_cols = list(pivot.columns)
    month_cols = [c for c in all_cols if c not in ["合计", "等通知", "未客审"]]
    month_cols_sorted = sorted(month_cols, key=get_sort_order)
    
    other_cols = []
    if "等通知" in all_cols:
        other_cols.append("等通知")
    if "未客审" in all_cols:
        other_cols.append("未客审")
    
    sorted_cols = month_cols_sorted + other_cols
    if "合计" in all_cols:
        sorted_cols.append("合计")
    
    existing_cols = [c for c in sorted_cols if c in pivot.columns]
    stat_table = pivot[existing_cols].copy()
    
    # 金额转为万元
    for col in stat_table.columns:
        stat_table[col] = stat_table[col] / 10000
    
    print(f"  统计表列顺序（已排序）: {list(stat_table.columns)}")
    
    # ========== 步骤8.5 按合计列降序排序，合计行保持最后 ==========
    print("\n【步骤8.5】按合计金额降序排序（合计行除外）...")
    
    # 分离合计行和数据行
    if "合计" in stat_table.index:
        total_row = stat_table.loc["合计"]
        data_rows = stat_table.drop("合计")
        
        if not data_rows.empty and "合计" in data_rows.columns:
            # 按合计列降序排序
            sorted_data_rows = data_rows.sort_values(by="合计", ascending=False)
            # 重新拼接
            stat_table = pd.concat([sorted_data_rows, total_row.to_frame().T])
    
    print(f"  排序后行顺序: {list(stat_table.index)}")
    
    # 验证5月金额
    may_col = None
    for col in stat_table.columns:
        if "05月" in str(col) or "5月" in str(col):
            may_col = col
            break
    
    if may_col and "合计" in stat_table.index:
        may_amount = stat_table.loc["合计", may_col]
        print(f"\n  ✅ 验证: {may_col} 金额 = {may_amount:.2f}万元")
    
    # ========== 步骤9: 保存Excel ==========
    print("\n【步骤9】保存Excel...")
    
    # 第一步：pandas保存
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sheet1_处理后', index=False)
        pivot.to_excel(writer, sheet_name='透视表')
        stat_table.to_excel(writer, sheet_name='统计表')
        
        wb = writer.book
        
        # 美化透视表：数值转为万元
        pivot_ws = writer.sheets['透视表']
        for row in pivot_ws.iter_rows(min_row=2, max_row=pivot_ws.max_row):
            for cell in row:
                if cell.column > 1 and isinstance(cell.value, (int, float)) and cell.value > 0:
                    cell.value = cell.value / 10000
    
    # 第二步：openpyxl美化统计表
    wb = load_workbook(OUTPUT_FILE)
    stat_ws = wb['统计表']
    
    # 插入标题行
    stat_ws.insert_rows(1)
    stat_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=stat_ws.max_column)
    title_cell = stat_ws.cell(row=1, column=1)
    title_cell.value = f'未发货订单统计 {today[:4]}年{today[4:6]}月{today[6:8]}日'
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    stat_ws.row_dimensions[1].height = 30
    
    # 样式定义
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置表头样式（第2行）
    for cell in stat_ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 修复：第一列（销售渠道）的列标题设为"销售渠道"
    stat_ws.cell(row=2, column=1).value = "销售渠道"
    
    # 设置数据样式
    for row in stat_ws.iter_rows(min_row=3, max_row=stat_ws.max_row):
        for cell in row:
            cell.border = border
            if cell.column > 1:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if isinstance(cell.value, (int, float)):
                    if cell.value == 0:
                        # 0值显示为空白
                        cell.value = None
                    else:
                        cell.number_format = '0.00"万"'
    
    # 第一列居左
    for row in stat_ws.iter_rows(min_row=3, max_row=stat_ws.max_row):
        row[0].alignment = Alignment(horizontal='left', vertical='center')
    
    # 自动调整列宽
    for column in stat_ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        adjusted_width = min(max(max_length + 2, 15), 35)
        stat_ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(OUTPUT_FILE)
    print(f"  Excel已保存: {OUTPUT_FILE}")
    
    print("\n" + "=" * 60)
    print("数据处理完成！")
    print("=" * 60)
    
    return OUTPUT_FILE


if __name__ == "__main__":
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
        output_dir = sys.argv[2] if len(sys.argv) > 2 else "."
        process_data(input_file, output_dir)
    else:
        print("使用方法: python process_data.py <输入文件> [输出目录]")
